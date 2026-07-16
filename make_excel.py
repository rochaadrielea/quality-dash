"""
make_excel.py
-------------
Exports BRM data to a multi-sheet Excel file for download.

Sheets:
    Summary          - Headline numbers (slide 2)
    Top6_Projects    - Top 6 projects (slides 3)
    By_Area          - NCs by detection area (slide 3)
    Prod_vs_Supplier - Production vs Supplier (slide 4)
    Monthly_Trend    - Opens vs closes by month (slide 4)
    Owner_Workload   - Open NCs by owner (weekly review)
    Full_Data        - All NCs from the DB (for filtering)

Usage:
    from make_excel import build_brm_excel
    xlsx_bytes = build_brm_excel()
"""

from io import BytesIO
import sqlite3

import pandas as pd
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

DB_FILE = "quality.db"

BG_NAVY_HEX = "1E2761"
BG_LIGHT_HEX = "F2F2F2"


def _q(sql):
    with sqlite3.connect(DB_FILE) as conn:
        return pd.read_sql(sql, conn)


def _format_sheet(ws, title):
    """Apply consistent header formatting."""
    # Column widths (autofit-ish)
    for col in ws.columns:
        max_len = 10
        letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value else 0
                if val_len > max_len:
                    max_len = val_len
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(max_len + 2, 45)

    # Header row styling
    header_fill = PatternFill("solid", fgColor=BG_NAVY_HEX)
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="left", vertical="center")

    ws.freeze_panes = "A2"


def build_brm_excel() -> bytes:
    """Build the multi-sheet BRM Excel. Returns .xlsx as bytes."""
    buf = BytesIO()

    # Gather all datasets
    datasets = {
        "Summary": _q("""
            SELECT
                (SELECT COUNT(*) FROM nc WHERE is_open=1) AS ncs_open_wip,
                (SELECT COUNT(*) FROM nc WHERE is_open=0) AS ncs_closed_total,
                (SELECT COUNT(*) FROM nc WHERE classification LIKE 'Major%' AND is_open=1) AS major_ncs_open,
                (SELECT COUNT(*) FROM nc WHERE is_supplier_nc=1 AND is_open=1) AS supplier_ncs_open,
                (SELECT COUNT(*) FROM nc WHERE is_supplier_nc=0 AND is_open=1) AS production_ncs_open,
                (SELECT COUNT(*) FROM nc WHERE project IS NULL) AS blank_project,
                (SELECT COUNT(*) FROM nc WHERE detection_area IS NULL) AS blank_detection
        """).T.reset_index().rename(columns={"index": "Metric", 0: "Value"}),

        "Top6_Projects_WIP": _q("""
            SELECT COALESCE(project, '(no project)') AS project, COUNT(*) AS open_ncs
            FROM nc WHERE is_open = 1
            GROUP BY project ORDER BY open_ncs DESC LIMIT 6
        """),

        "Top6_Projects_YTD": _q("""
            SELECT COALESCE(project, '(no project)') AS project, COUNT(*) AS ncs_opened_ytd
            FROM nc WHERE substr(created_on, 1, 4) = '2026'
            GROUP BY project ORDER BY ncs_opened_ytd DESC LIMIT 6
        """),

        "By_Area": _q("""
            SELECT COALESCE(detection_area, 'BLANK - to clean') AS area, COUNT(*) AS ncs
            FROM nc WHERE substr(created_on, 1, 4) = '2026'
            GROUP BY area ORDER BY ncs DESC
        """),

        "Prod_vs_Supplier": _q("""
            SELECT
                CASE WHEN is_supplier_nc = 1 THEN 'Supplier' ELSE 'Production' END AS source,
                SUM(CASE WHEN is_open = 1 THEN 1 ELSE 0 END) AS open_ncs,
                SUM(CASE WHEN is_open = 0 THEN 1 ELSE 0 END) AS closed_ncs
            FROM nc GROUP BY source
        """),

        "Monthly_Trend": _q("""
            WITH opens AS (
                SELECT substr(created_on, 1, 7) AS month, COUNT(*) AS opened
                FROM nc WHERE created_on IS NOT NULL GROUP BY month
            ),
            closes AS (
                SELECT substr(closure_date, 1, 7) AS month, COUNT(*) AS closed
                FROM nc WHERE closure_date IS NOT NULL GROUP BY month
            )
            SELECT o.month, o.opened, COALESCE(c.closed, 0) AS closed
            FROM opens o LEFT JOIN closes c USING (month)
            ORDER BY o.month
        """),

        "Owner_Workload": _q("""
            SELECT COALESCE(owner, '(no owner)') AS owner,
                   SUM(CASE WHEN is_open = 1 THEN 1 ELSE 0 END) AS open_ncs,
                   SUM(CASE WHEN is_open = 0 THEN 1 ELSE 0 END) AS closed_ncs,
                   MAX(days_open) AS oldest_days
            FROM nc GROUP BY owner ORDER BY open_ncs DESC
        """),

        "Data_Quality_Check": _q("""
            SELECT nc_id, project, detection_area, classification, owner, status
            FROM nc
            WHERE project IS NULL
               OR detection_area IS NULL
               OR classification IS NULL
               OR owner IS NULL
            ORDER BY nc_id
        """),

        "Full_Data": _q("SELECT * FROM nc ORDER BY created_on DESC"),
    }

    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        for sheet_name, df in datasets.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

        # apply formatting after write
        for sheet_name in datasets:
            _format_sheet(writer.sheets[sheet_name], sheet_name)

    buf.seek(0)
    return buf.read()


if __name__ == "__main__":
    data = build_brm_excel()
    with open("BRM_Quality_June_2026.xlsx", "wb") as f:
        f.write(data)
    print(f"✓ Written BRM_Quality_June_2026.xlsx ({len(data)/1024:.1f} KB)")
